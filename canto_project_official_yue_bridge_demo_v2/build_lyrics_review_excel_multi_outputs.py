#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Build an Excel review table that simultaneously aggregates lyrics from
multiple output directories.

This script is designed for comparing different runs / output folders, for example:

outputs/batch_eval_preset/
  intern/
    json/*.json
  intern_rag/
    json/*.json

outputs/batch_eval_image_mood/
  intern/
    json/*.json
  intern_rag/
    json/*.json

The generated Excel has one row per image, with the original image thumbnail and
lyrics columns from every selected output directory + model combination.

Usage example: compare different output dirs, each containing model subfolders

python build_lyrics_review_excel_multi_outputs.py \
  --image-dir /userhome/cs5/u3665806/Image2CantonSong/Images/flickr8k/Images \
  --output-dirs outputs/batch_eval_preset outputs/batch_eval_image_mood \
  --run-labels preset image_mood \
  --models intern_rag intern \
  --output-xlsx outputs/lyrics_review_multi_outputs.xlsx \
  --max-images 50 \
  --include-genre \
  --include-status \
  --include-json-path

This creates columns like:
preset_intern_rag_title
preset_intern_rag_lyrics
preset_intern_title
preset_intern_lyrics
image_mood_intern_rag_title
image_mood_intern_rag_lyrics
image_mood_intern_title
image_mood_intern_lyrics

Alternative usage: directly provide source dirs

python build_lyrics_review_excel_multi_outputs.py \
  --image-dir /path/to/images \
  --source-dirs outputs/run1/intern_rag outputs/run2/intern_rag outputs/run1/intern \
  --source-labels run1_rag run2_rag run1_intern \
  --output-xlsx outputs/lyrics_review_direct_sources.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from PIL import Image as PILImage
except ImportError as exc:
    raise ImportError("Missing Pillow. Install with: pip install pillow") from exc

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError as exc:
    raise ImportError("Missing openpyxl. Install with: pip install openpyxl") from exc


SUPPORTED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create an Excel review table with original images and lyrics "
            "from multiple output directories."
        )
    )

    parser.add_argument(
        "--image-dir",
        type=Path,
        required=True,
        help="Directory containing original images.",
    )

    parser.add_argument(
        "--output-dirs",
        nargs="+",
        type=Path,
        default=None,
        help=(
            "Parent output directories. Each parent should contain model folders, "
            "for example outputs/batch_eval_preset/intern_rag/json/*.json."
        ),
    )

    parser.add_argument(
        "--run-labels",
        nargs="+",
        default=None,
        help=(
            "Optional labels for --output-dirs. Must match the number of output dirs. "
            "Example: --run-labels preset image_mood"
        ),
    )

    parser.add_argument(
        "--models",
        nargs="+",
        default=None,
        help=(
            "Model folder names under every --output-dirs parent. "
            "Example: --models intern_rag intern qwen"
        ),
    )

    parser.add_argument(
        "--source-dirs",
        nargs="+",
        type=Path,
        default=None,
        help=(
            "Direct source directories. Use this if you want full control. "
            "Each source directory should contain json/*.json or metrics_json/*.json."
        ),
    )

    parser.add_argument(
        "--source-labels",
        nargs="+",
        default=None,
        help="Labels for --source-dirs. Must match the number of source dirs.",
    )

    parser.add_argument(
        "--output-xlsx",
        type=Path,
        required=True,
        help="Path to save Excel file.",
    )

    parser.add_argument(
        "--start-index",
        type=int,
        default=0,
        help="Start image index, 0-based.",
    )

    parser.add_argument(
        "--max-images",
        type=int,
        default=0,
        help="Maximum number of images to include. 0 means no limit.",
    )

    parser.add_argument(
        "--image-width",
        type=int,
        default=180,
        help="Thumbnail width in pixels.",
    )

    parser.add_argument(
        "--image-height",
        type=int,
        default=120,
        help="Maximum thumbnail height in pixels.",
    )

    parser.add_argument(
        "--include-genre",
        action="store_true",
        help="Include genre_prompt columns.",
    )

    parser.add_argument(
        "--include-status",
        action="store_true",
        help="Include status columns from metrics JSON if available.",
    )

    parser.add_argument(
        "--include-json-path",
        action="store_true",
        help="Include JSON path columns.",
    )

    return parser.parse_args()


def safe_label(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r"[^0-9A-Za-z_\-]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_") or "source"


def clean_cell_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text


def list_images(image_dir: Path) -> list[Path]:
    if not image_dir.exists():
        raise FileNotFoundError(f"Image directory not found: {image_dir}")

    if not image_dir.is_dir():
        raise NotADirectoryError(f"Image path is not a directory: {image_dir}")

    return sorted(
        p for p in image_dir.iterdir()
        if p.is_file() and p.suffix.lower() in SUPPORTED_IMAGE_EXTS
    )


def resolve_sources(args: argparse.Namespace) -> list[dict[str, Any]]:
    """
    Return list of sources:
    {
      "label": str,
      "dir": Path,
      "run_label": str,
      "model": str
    }

    Two modes:

    Mode A:
      --output-dirs run1 run2 --run-labels preset image_mood --models intern_rag intern

    Mode B:
      --source-dirs dir1 dir2 --source-labels label1 label2
    """

    sources: list[dict[str, Any]] = []

    if args.source_dirs:
        if args.output_dirs:
            raise ValueError("Use either --source-dirs or --output-dirs, not both.")

        if args.source_labels and len(args.source_labels) != len(args.source_dirs):
            raise ValueError("--source-labels length must match --source-dirs length.")

        for i, source_dir in enumerate(args.source_dirs):
            label = args.source_labels[i] if args.source_labels else source_dir.name
            sources.append(
                {
                    "label": safe_label(label),
                    "dir": source_dir,
                    "run_label": "",
                    "model": "",
                }
            )

    else:
        if not args.output_dirs:
            raise ValueError("Please provide either --output-dirs or --source-dirs.")

        if not args.models:
            raise ValueError("When using --output-dirs, you must provide --models.")

        if args.run_labels and len(args.run_labels) != len(args.output_dirs):
            raise ValueError("--run-labels length must match --output-dirs length.")

        for i, output_dir in enumerate(args.output_dirs):
            run_label = args.run_labels[i] if args.run_labels else output_dir.name
            run_label = safe_label(run_label)

            for model in args.models:
                model_label = safe_label(model)
                source_dir = output_dir / model
                combined_label = safe_label(f"{run_label}_{model_label}")

                sources.append(
                    {
                        "label": combined_label,
                        "dir": source_dir,
                        "run_label": run_label,
                        "model": model_label,
                    }
                )

    # De-duplicate labels if needed
    seen: dict[str, int] = {}
    for source in sources:
        base = source["label"]
        if base not in seen:
            seen[base] = 1
        else:
            seen[base] += 1
            source["label"] = f"{base}_{seen[base]}"

    for source in sources:
        if not source["dir"].exists():
            print(f"Warning: source dir does not exist: {source['dir']}", file=sys.stderr)

    return sources


def load_json(path: Path) -> dict[str, Any] | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def find_payload(source_dir: Path, image_stem: str) -> tuple[dict[str, Any] | None, Path | None]:
    """
    Find generated JSON for a given image stem.

    Priority:
    1. source_dir/json/stem.json
    2. source_dir/metrics_json/stem.json
    3. source_dir/stem.json
    4. recursive fallback
    """

    candidates = [
        source_dir / "json" / f"{image_stem}.json",
        source_dir / "metrics_json" / f"{image_stem}.json",
        source_dir / f"{image_stem}.json",
    ]

    for path in candidates:
        if path.exists():
            payload = load_json(path)
            if payload is not None:
                return payload, path

    if source_dir.exists():
        matches = list(source_dir.rglob(f"{image_stem}.json"))

        matches = sorted(
            matches,
            key=lambda p: (
                0 if p.parent.name == "json" else 1 if p.parent.name == "metrics_json" else 2,
                len(str(p)),
            ),
        )

        for path in matches:
            payload = load_json(path)
            if payload is not None:
                return payload, path

    return None, None


def extract_lyrics_fields(payload: dict[str, Any] | None) -> dict[str, str]:
    if not payload:
        return {
            "title": "",
            "lyrics_text": "",
            "genre_prompt": "",
            "status": "missing",
        }

    return {
        "title": clean_cell_text(payload.get("title", "")),
        "lyrics_text": clean_cell_text(payload.get("lyrics_text", "")),
        "genre_prompt": clean_cell_text(payload.get("genre_prompt", "")),
        "status": clean_cell_text(payload.get("status", "")),
    }


def add_thumbnail(ws, image_path: Path, cell: str, max_width: int, max_height: int) -> None:
    try:
        with PILImage.open(image_path) as img:
            width, height = img.size

        if width <= 0 or height <= 0:
            return

        ratio = min(max_width / width, max_height / height)
        display_width = max(1, int(width * ratio))
        display_height = max(1, int(height * ratio))

        xl_img = XLImage(str(image_path))
        xl_img.width = display_width
        xl_img.height = display_height
        ws.add_image(xl_img, cell)

    except Exception as exc:
        print(f"Warning: failed to add image {image_path}: {exc}", file=sys.stderr)


def set_basic_styles(ws, total_cols: int, frozen_cell: str = "E2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = frozen_cell
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def build_headers(
    sources: list[dict[str, Any]],
    include_genre: bool,
    include_status: bool,
    include_json_path: bool,
) -> list[str]:
    headers = ["idx", "image_name", "original_image", "image_path"]

    for source in sources:
        label = source["label"]
        headers.extend(
            [
                f"{label}_title",
                f"{label}_lyrics",
            ]
        )

        if include_genre:
            headers.append(f"{label}_genre_prompt")

        if include_status:
            headers.append(f"{label}_status")

        if include_json_path:
            headers.append(f"{label}_json_path")

    return headers


def build_workbook(
    images: list[Path],
    sources: list[dict[str, Any]],
    output_xlsx: Path,
    image_width: int,
    image_height: int,
    include_genre: bool,
    include_status: bool,
    include_json_path: bool,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lyrics Review"

    headers = build_headers(
        sources=sources,
        include_genre=include_genre,
        include_status=include_status,
        include_json_path=include_json_path,
    )
    ws.append(headers)

    source_stats = {
        source["label"]: {
            "found": 0,
            "missing": 0,
            "source_dir": str(source["dir"]),
            "run_label": source.get("run_label", ""),
            "model": source.get("model", ""),
        }
        for source in sources
    }

    for row_idx, image_path in enumerate(images, start=2):
        excel_idx = row_idx - 1
        image_stem = image_path.stem

        row_values: list[Any] = [
            excel_idx,
            image_path.name,
            "",
            str(image_path),
        ]

        for source in sources:
            payload, json_path = find_payload(source["dir"], image_stem)
            fields = extract_lyrics_fields(payload)

            if payload is None:
                source_stats[source["label"]]["missing"] += 1
            else:
                source_stats[source["label"]]["found"] += 1

            row_values.extend(
                [
                    fields["title"],
                    fields["lyrics_text"],
                ]
            )

            if include_genre:
                row_values.append(fields["genre_prompt"])

            if include_status:
                row_values.append(fields["status"])

            if include_json_path:
                row_values.append(str(json_path) if json_path else "")

        ws.append(row_values)

        add_thumbnail(
            ws=ws,
            image_path=image_path,
            cell=f"C{row_idx}",
            max_width=image_width,
            max_height=image_height,
        )

        ws.row_dimensions[row_idx].height = max(90, int(image_height * 0.75) + 12)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = max(18, image_width / 7)
    ws.column_dimensions["D"].width = 45

    for col_idx in range(5, len(headers) + 1):
        header = headers[col_idx - 1]
        col_letter = get_column_letter(col_idx)

        if header.endswith("_lyrics"):
            ws.column_dimensions[col_letter].width = 46
        elif header.endswith("_genre_prompt"):
            ws.column_dimensions[col_letter].width = 36
        elif header.endswith("_json_path"):
            ws.column_dimensions[col_letter].width = 48
        else:
            ws.column_dimensions[col_letter].width = 24

    set_basic_styles(ws, total_cols=len(headers), frozen_cell="E2")

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["item", "value"])
    summary_ws.append(["image_count", len(images)])
    summary_ws.append(["source_count", len(sources)])
    summary_ws.append(["output_xlsx", str(output_xlsx)])
    summary_ws.append([])

    summary_ws.append(["source_label", "run_label", "model", "source_dir", "found_json", "missing_json"])
    for source in sources:
        label = source["label"]
        stat = source_stats[label]
        summary_ws.append(
            [
                label,
                stat["run_label"],
                stat["model"],
                stat["source_dir"],
                stat["found"],
                stat["missing"],
            ]
        )

    set_basic_styles(summary_ws, total_cols=6, frozen_cell="A2")
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 22
    summary_ws.column_dimensions["C"].width = 18
    summary_ws.column_dimensions["D"].width = 75
    summary_ws.column_dimensions["E"].width = 16
    summary_ws.column_dimensions["F"].width = 16

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    args = parse_args()

    all_images = list_images(args.image_dir)

    start_index = max(0, int(args.start_index))
    images = all_images[start_index:]

    if args.max_images > 0:
        images = images[: args.max_images]

    if not images:
        raise ValueError("No images selected. Check --image-dir, --start-index, and --max-images.")

    sources = resolve_sources(args)

    print(f"Total images found: {len(all_images)}")
    print(f"Images selected: {len(images)}")
    print("Sources:")
    for source in sources:
        print(f"  - {source['label']}: {source['dir']}")

    build_workbook(
        images=images,
        sources=sources,
        output_xlsx=args.output_xlsx,
        image_width=args.image_width,
        image_height=args.image_height,
        include_genre=args.include_genre,
        include_status=args.include_status,
        include_json_path=args.include_json_path,
    )

    print(f"Saved Excel review file to: {args.output_xlsx}")


if __name__ == "__main__":
    main()
